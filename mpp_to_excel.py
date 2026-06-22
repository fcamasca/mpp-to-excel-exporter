"""Exporta tareas de Microsoft Project (.mpp) a un Excel legible.

MPXJ realiza la lectura del formato propietario y entrega un JSON temporal;
este módulo transforma ese JSON y aplica formato al libro de salida.
"""

from __future__ import annotations

import argparse
import json
import os
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill


COLUMNS = [
    "Nombre de tarea",
    "Duración",
    "Trabajo",
    "Comienzo",
    "Fin",
    "% completado",
    "Nombres de los recursos",
]


def parse_date(value: str | None) -> datetime | None:
    """Convierte una fecha ISO de MPXJ en ``datetime``."""
    if not value:
        return None
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return None


def format_duration(seconds: float | int | None) -> str:
    """Presenta segundos como minutos, horas o jornadas de ocho horas."""
    minutes = (seconds or 0) / 60
    if minutes <= 0:
        return "0 días"
    if minutes < 60:
        return f"{int(minutes)} min"
    if minutes < 480:
        hours = minutes / 60
        return "1 hora" if hours == 1 else f"{hours:g} horas"
    days = minutes / 480
    return "1 día" if days == 1 else f"{days:g} días"


def extract_rows(data: dict[str, Any]) -> tuple[list[dict[str, Any]], list[bool]]:
    """Extrae las filas de tareas y las marcas de tareas resumen."""
    resources = {
        resource.get("unique_id"): resource.get("name")
        for resource in data.get("resources", [])
        if resource.get("unique_id") and resource.get("name")
    }

    assignments: dict[Any, list[Any]] = {}
    for assignment in data.get("assignments", []):
        task_id = assignment.get("task_unique_id")
        resource_id = assignment.get("resource_unique_id")
        if task_id and resource_id:
            assignments.setdefault(task_id, []).append(resource_id)

    rows: list[dict[str, Any]] = []
    summary_flags: list[bool] = []
    for task in data.get("tasks", []):
        name = task.get("name")
        if not name:
            continue

        level = max(int(task.get("outline_level") or 1), 1)
        task_resources = {
            resources[resource_id]
            for resource_id in assignments.get(task.get("unique_id"), [])
            if resource_id in resources
        }
        percent = float(task.get("percent_complete") or 0) / 100

        rows.append(
            {
                "Nombre de tarea": "      " * (level - 1) + name,
                "Duración": format_duration(task.get("duration")),
                "Trabajo": float(task.get("work") or 0) / 3600,
                "Comienzo": parse_date(task.get("start")),
                "Fin": parse_date(task.get("finish")),
                "% completado": percent,
                "Nombres de los recursos": ", ".join(sorted(task_resources)),
            }
        )
        summary_flags.append(bool(task.get("summary", False)))

    return rows, summary_flags


def write_excel(data: dict[str, Any], output_file: Path) -> int:
    """Crea el Excel y devuelve la cantidad de tareas exportadas."""
    rows, summary_flags = extract_rows(data)
    frame = pd.DataFrame(rows, columns=COLUMNS)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        frame.to_excel(writer, index=False, sheet_name="Tareas")
        sheet = writer.sheets["Tareas"]
        widths = {"A": 60, "B": 12, "C": 10, "D": 14, "E": 14, "F": 12, "G": 30}
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width

        header_fill = PatternFill("solid", fgColor="D9D9D9")
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="left", vertical="center")

        for index, excel_row in enumerate(range(2, len(frame) + 2)):
            name = str(sheet.cell(excel_row, 1).value or "")
            font = None
            if summary_flags[index]:
                font = Font(bold=True)
            elif "Hito:" in name:
                font = Font(color="0000FF", italic=True)
            if font:
                for column in range(1, len(COLUMNS) + 1):
                    sheet.cell(excel_row, column).font = font

            for column in (4, 5):
                sheet.cell(excel_row, column).number_format = "DDD DD/MM/YY"
            sheet.cell(excel_row, 6).number_format = "0%"
            sheet.cell(excel_row, 6).alignment = Alignment(horizontal="center")

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

    return len(frame)


def build_mpxj_classpath(mpxj_home: Path) -> str:
    """Construye el classpath sin depender de scripts sensibles a espacios."""
    mpxj_jar = mpxj_home / "mpxj.jar"
    library_dir = mpxj_home / "lib"
    if not mpxj_jar.is_file() or not library_dir.is_dir():
        raise FileNotFoundError(
            "La distribución de MPXJ debe contener mpxj.jar y la carpeta lib: "
            f"{mpxj_home}"
        )
    return os.pathsep.join((str(library_dir / "*"), str(mpxj_jar)))


def convert_to_json(input_file: Path, json_file: Path, mpxj_home: Path) -> None:
    """Invoca la utilidad oficial de MPXJ sin usar ``shell=True``."""
    command = [
        "java",
        "-cp",
        build_mpxj_classpath(mpxj_home),
        "org.mpxj.sample.MpxjConvert",
        str(input_file),
        str(json_file),
    ]
    result = subprocess.run(command, check=False, capture_output=True, text=True)
    if result.returncode:
        detail = (result.stderr or result.stdout).strip()
        raise RuntimeError(f"MPXJ no pudo convertir el archivo. {detail}")


def convert_file(input_file: Path, output_file: Path, mpxj_home: Path) -> int:
    if not input_file.is_file():
        raise FileNotFoundError(f"No existe el archivo: {input_file}")
    json_file = input_file.with_suffix(input_file.suffix + ".temp.json")
    try:
        convert_to_json(input_file, json_file, mpxj_home)
        with json_file.open(encoding="utf-8") as stream:
            return write_excel(json.load(stream), output_file)
    finally:
        json_file.unlink(missing_ok=True)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Exporta tareas de archivos MPP a Excel.")
    parser.add_argument("input", type=Path, help="Archivo .mpp de entrada")
    parser.add_argument("-o", "--output", type=Path, help="Excel de salida (por defecto, mismo nombre)")
    parser.add_argument(
        "--mpxj-home",
        type=Path,
        default=Path(os.environ.get("MPXJ_HOME", "mpxj")),
        help="Carpeta de MPXJ (o variable MPXJ_HOME)",
    )
    return parser


def main() -> int:
    args = build_parser().parse_args()
    output = args.output or args.input.with_suffix(".xlsx")
    try:
        count = convert_file(args.input.resolve(), output.resolve(), args.mpxj_home.resolve())
    except (FileNotFoundError, RuntimeError, json.JSONDecodeError) as error:
        print(f"Error: {error}", file=sys.stderr)
        return 1
    print(f"Exportación completada: {count} tareas → {output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
